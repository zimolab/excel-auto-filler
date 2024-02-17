import dataclasses
import os.path
import webbrowser

from PySide2.QtGui import QTextCursor
from PySide2.QtWidgets import QMainWindow, QMessageBox, QFileDialog
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from excelfiller.rule import CellRulesConfig
from excelfiller_executable.common import DEFAULT_ENCODING, create_xlsx_file
from excelfiller_executable.gui.generated_code.main_window import Ui_MainWindow
from excelfiller_executable.gui.ui.common import RULES_CONFIG_FILE_FILTER, PROG_NAME, VERSION, WEBSITE
from excelfiller_executable.gui.ui.convert_dialog import ConvertDialog
from excelfiller_executable.gui.ui.mainwindow import terminal_styles
from excelfiller_executable.gui.ui.mainwindow.about_dialog import AboutDialog
from excelfiller_executable.gui.ui.mainwindow.controller import CellRulesProcessingController
from excelfiller_executable.gui.ui.mainwindow.handler import _ProcessorEventHandler
from excelfiller_executable.gui.ui.mainwindow.terminal_styles import LogLevel
from excelfiller_executable.gui.ui.new_rules_dialog import NewRulesConfigDialog

WINDOW_TITLE = f"{PROG_NAME} V{VERSION}"


@dataclasses.dataclass
class _Arguments(object):
    target_file: str
    rules_config_file: str
    sheet: int | str = ""
    save_as_file: str = ""
    encoding: str = DEFAULT_ENCODING
    overwrite: bool = False
    create: bool = False
    quick_fail: bool = False
    save_anyway: bool = False


class RuleProcessorWindow(QMainWindow):
    def __init__(self):
        super().__init__(parent=None)
        self._ui = Ui_MainWindow()
        self._processing_controller = CellRulesProcessingController(parent=self)
        self._event_handler = _ProcessorEventHandler(self)

        self.setup_ui()
        self.setup_signals()

        self._event_handler.takeover(self)

    def setup_ui(self):
        self._ui.setupUi(self)
        self.setWindowTitle(WINDOW_TITLE)
        self.setup_output_textbrowser()

        self.update_ui_with_arguments(_Arguments(
            target_file="",
            sheet="",
            rules_config_file="",
            save_as_file="",
            encoding=DEFAULT_ENCODING,
            overwrite=False,
            create=False,
            quick_fail=False,
            save_anyway=False,
        ))

    def setup_signals(self):
        self._ui.button_select_target_file.clicked.connect(self.select_target_file)
        self._ui.button_select_rules_config_file.clicked.connect(self.select_rules_config_file)
        self._ui.button_start.clicked.connect(self.start_processing)
        self._ui.button_select_save_as_file.clicked.connect(self.select_save_as_file)
        self._ui.button_stop.clicked.connect(self.stop_processing)
        self._ui.button_clear_output.clicked.connect(self.clear_output)
        self._ui.button_export_output.clicked.connect(self.export_output_to_file)
        self._ui.button_quit.clicked.connect(self.close)

        self._ui.action_convert.triggered.connect(self.on_action_convert_triggered)
        self._ui.action_new_rules_config.triggered.connect(self.on_action_new_rules_config_triggered)
        self._ui.action_about.triggered.connect(self.open_about_dialog)
        self._ui.action_help.triggered.connect(self.open_help_dialog)

    def on_action_convert_triggered(self):
        dialog = ConvertDialog(self)
        dialog.exec_()

    def on_action_new_rules_config_triggered(self):
        dialog = NewRulesConfigDialog(self)
        dialog.exec_()

    def setup_output_textbrowser(self):
        self._ui.textbrowser_output.setOpenExternalLinks(False)
        self._ui.textbrowser_output.setAcceptRichText(True)
        self._ui.textbrowser_output.setReadOnly(True)

        self._ui.textbrowser_output.setStyleSheet(terminal_styles.stylesheet())

    def collect_arguments(self) -> _Arguments:
        arguments = _Arguments(
            target_file=self._ui.edit_target_file.text().strip(),
            sheet=self._ui.edit_sheetname.text(),
            rules_config_file=self._ui.edit_rules_config_file.text().strip(),
            save_as_file=self._ui.edit_save_as_file.text().strip() or "",
            encoding=self._ui.edit_encoding.text().strip() or DEFAULT_ENCODING,
            overwrite=self._ui.check_overwrite.isChecked(),
            create=self._ui.checkbox_create.isChecked(),
            quick_fail=self._ui.checkbox_quick_fail.isChecked(),
            save_anyway=self._ui.checkbox_save_anyway.isChecked(),
        )
        return arguments

    def validate_arguments(self, arguments: _Arguments) -> bool:
        if not arguments.target_file:
            self.show_error("请填写目标文件路径")
            return False
        if not arguments.rules_config_file:
            self.show_error("请填写规则配置文件路径")
            return False

        if not os.path.isfile(arguments.rules_config_file):
            self.show_error("规则配置文件不存在")
            return False
        save_as_file = arguments.save_as_file
        if save_as_file:
            save_path = os.path.normcase(os.path.abspath(save_as_file))
        else:
            save_path = os.path.normcase(os.path.abspath(arguments.target_file))

        if not arguments.overwrite and os.path.isfile(save_path):
            self.show_error("文件保存路径已存在文件，如需覆盖，请先选择【允许覆盖已存在的文件】选项！")
            return False

        return True

    def update_ui_with_arguments(self, arguments: _Arguments):
        self._ui.edit_target_file.setText(arguments.target_file)
        self._ui.edit_sheetname.setText(arguments.sheet)
        self._ui.edit_rules_config_file.setText(arguments.rules_config_file)
        self._ui.edit_save_as_file.setText(arguments.save_as_file or "")
        self._ui.edit_encoding.setText(arguments.encoding)
        self._ui.check_overwrite.setChecked(arguments.overwrite)
        self._ui.checkbox_create.setChecked(arguments.create)
        self._ui.checkbox_quick_fail.setChecked(arguments.quick_fail)
        self._ui.checkbox_save_anyway.setChecked(arguments.save_anyway)

    def select_target_file(self):
        selected_file, _ = QFileDialog.getOpenFileName(self, "选择目标文件", filter="Excel工作簿文件(*.xlsx)")
        if not selected_file:
            return
        self._ui.edit_target_file.setText(os.path.normcase(os.path.abspath(selected_file)))

    def select_rules_config_file(self):
        selected_file, _ = QFileDialog.getOpenFileName(self, "选择规则配置文件", filter=RULES_CONFIG_FILE_FILTER)
        if not selected_file:
            return
        self._ui.edit_rules_config_file.setText(os.path.normcase(os.path.abspath(selected_file)))

    def start_processing(self):
        # 获取当前参数
        arguments = self.collect_arguments()
        # 校验当前参数
        if not self.validate_arguments(arguments):
            return
        if self._processing_controller.processor.is_processing():
            self.show_error("正在处理中，请等待处理完成！")
            return
        rules_config = self.load_rules_config(arguments.rules_config_file, arguments.encoding)
        if not rules_config:
            return
        workbook = self.get_workbook(arguments.target_file, arguments.create)
        if not workbook:
            return
        worksheet = self.get_worksheet(workbook, arguments.sheet)
        if not worksheet:
            return

        try:
            save_as_file = arguments.save_as_file
            if save_as_file:
                save_path = os.path.normcase(os.path.abspath(save_as_file))
            else:
                save_path = os.path.normcase(os.path.abspath(arguments.target_file))
        except BaseException as e:
            self.show_error(f"另存为文件路径错误：{e}")
            return

        if self._processing_controller.is_processing():
            self.show_warning("正在处理中，请等待处理完成！")
            return
        self._ui.button_start.setEnabled(False)
        self._processing_controller.start_processing(
            workbook=workbook,
            worksheet=worksheet,
            rules_config=rules_config,
            save_path=save_path,
            save_anyway=arguments.save_anyway,
            overwrite=arguments.overwrite,
            quick_fail=arguments.quick_fail,
        )

    def stop_processing(self):
        if not self._processing_controller.is_processing():
            return
        self._processing_controller.stop()

    def get_workbook(self, path: str, create: bool) -> Workbook | None:
        if create and not os.path.isfile(path):
            workbook = create_xlsx_file(path=path)
            if os.path.isfile(path):
                return workbook
            else:
                self.show_error("创建文件失败!")
                return None

        if os.path.isfile(path):
            try:
                workbook = load_workbook(path)
            except BaseException as e:
                self.show_error(f"加载文件失败：{e}")
                return None
            else:
                return workbook
        else:
            self.show_error(f"未找到文件：'{path}'")
            return None

    def get_worksheet(self, workbook: Workbook, sheet: str | int | None) -> Worksheet | None:
        if not sheet:
            return workbook.active

        if isinstance(sheet, str):
            if sheet in workbook.sheetnames:
                return workbook[sheet]
            try:
                worksheet = workbook.create_sheet(title=sheet)
            except BaseException as e:
                self.show_error(f"创建工作表失败：{e}")
                return None
            else:
                return worksheet

        try:
            worksheet = workbook.worksheets[sheet]
        except BaseException as e:
            self.show_error(f"获取工作表失败：{e}")
            return None
        else:
            return worksheet

    def load_rules_config(self, path: str, encoding: str) -> CellRulesConfig | None:
        # 加载规则配置文件
        try:
            rules_config = self._processing_controller.config_factory.load(path=path, encoding=encoding)
        except BaseException as e:
            self.show_error(f"加载规则配置文件失败：{e}")
            return None
        else:
            return rules_config

    def select_save_as_file(self):
        selected_file, _ = QFileDialog.getSaveFileName(self, "另存为", filter="Excel工作簿文件(*.xlsx)")
        if not selected_file:
            return
        selected_file = os.path.normcase(os.path.abspath(selected_file))
        if os.path.isfile(selected_file):
            if not self.show_confirm("文件已存在，是否覆盖？如果确认覆盖，【允许覆盖已存在的文件】选项将被自动勾选",
                                     title="确认覆盖？"):
                return
            self._ui.check_overwrite.setChecked(True)
            self._ui.edit_save_as_file.setText(selected_file)
        else:
            self._ui.edit_save_as_file.setText(selected_file)

    def clear_output(self):
        self._ui.textbrowser_output.clear()

    def export_output_to_file(self):
        selected_file, _ = QFileDialog.getSaveFileName(self, "导出输出到文件", filter="文本文件(*.txt);所有文件(*.*)")
        if not selected_file:
            return
        content = self._ui.textbrowser_output.toPlainText()
        try:
            with open(selected_file, "w", encoding="utf-8") as f:
                f.write(content)
        except BaseException as e:
            self.show_error(f"导出输出到文件失败：{e}")
        else:
            self.show_info(f"导出输出到文件成功：{selected_file}")

    def output(self, message: str):
        cursor: QTextCursor = self._ui.textbrowser_output.textCursor()
        cursor.insertBlock()
        cursor.insertHtml(message)
        self._ui.textbrowser_output.ensureCursorVisible()
        self._ui.textbrowser_output.moveCursor(cursor.End)

    def output_info(self, msg: str, timestamp: bool = True, additional_nl: bool = False):
        self.output(terminal_styles.log(LogLevel.INFO, msg, timestamp=timestamp, additional_nl=additional_nl))

    def output_debug(self, msg: str, timestamp: bool = True, additional_nl: bool = False):
        self.output(terminal_styles.log(LogLevel.DEBUG, msg, timestamp=timestamp, additional_nl=additional_nl))

    def output_warning(self, msg: str, timestamp: bool = True, additional_nl: bool = False):
        self.output(terminal_styles.log(LogLevel.WARNING, msg, timestamp=timestamp, additional_nl=additional_nl))

    def output_critical(self, msg: str, timestamp: bool = True, additional_nl: bool = False):
        self.output(terminal_styles.log(LogLevel.CRITICAL, msg, timestamp=timestamp, additional_nl=additional_nl))

    def output_fatal(self, msg: str, timestamp: bool = True, additional_nl: bool = False):
        self.output(terminal_styles.log(LogLevel.FATAL, msg, timestamp=timestamp, additional_nl=additional_nl))

    def is_clear_output_before_start(self) -> bool:
        return self._ui.checkbox_clear_output.isChecked()

    def show_info(self, message: str, title: str = "信息", *args):
        return QMessageBox.information(self, title, message, *args)

    def show_error(self, message: str, title: str = "错误", *args):
        return QMessageBox.critical(self, title, message, *args)

    def show_warning(self, message: str, title: str = "警告", *args):
        return QMessageBox.warning(self, title, message, *args)

    def show_confirm(self, message: str, title: str = "确认"):
        return QMessageBox.question(self, title, message, QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes

    def set_start_button_state(self, enabled: bool):
        self._ui.button_start.setEnabled(enabled)

    def closeEvent(self, event):
        if self._processing_controller.is_processing():
            self.show_warning("任务正在处理中，无法关闭窗口！")
            event.ignore()
            return
        event.accept()

    @property
    def processing_controller(self):
        return self._processing_controller

    @property
    def ui(self):
        return self._ui

    def open_about_dialog(self):
        dialog = AboutDialog(self)
        dialog.exec_()

    def open_help_dialog(self):
        if self.show_confirm("目前帮助文档尚未编写，本项目为开源项目，所有源码均已公开，是否跳转到本项目仓库？"):
            webbrowser.open(WEBSITE)
