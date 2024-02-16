import os
from typing import List, Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from excelfiller.common import PrimitiveDataType, is_primitive_data_type
from excelfiller.function import FunctionContext
from excelfiller.processor import CellRulesProcessor
from excelfiller.rule import CellRulesConfig, CellRule
from excelfiller_executable.cli.cli import CLICommand, CellRulesProcessorCLI, STYLE_INFO, STYLE_WARN, STYLE_SUCCESS, \
    STYLE_HINT, STYLE_ERROR
from excelfiller_executable.common import create_xlsx_file, DEFAULT_ENCODING


# noinspection PyUnusedLocal
class CommandRunRules(CLICommand):
    def __init__(self, cli: "CellRulesProcessorCLI", config_file: str, config_file_encoding: str, target_xlsx: str,
                 sheet: str | None = None, create: bool = False, save_as: str | None = None, run_async: bool = False,
                 quick_fail: bool = False, overwrite: bool = False, save_anyway: bool = False, verbose: bool = False):
        super().__init__(cli)
        self.config_file = config_file
        self.config_file_encoding = config_file_encoding or DEFAULT_ENCODING
        self.target_xlsx = target_xlsx
        self.sheet = sheet
        self.create = create
        self.save_as = save_as or None
        self.run_async = run_async
        self.quick_fail = quick_fail
        self.overwrite = overwrite
        self.save_anyway = save_anyway
        self.verbose = verbose

        self._execution_err = None

    def _on_started(self, processor: CellRulesProcessor, workbook: Workbook, worksheet: Worksheet,
                    rules_config: CellRulesConfig):
        self.cli.print("[Info] Processing started!", style=STYLE_INFO)
        self.display_rules_config_info(rules_config)

    def _on_finished(self, processor: CellRulesProcessor, workbook: Workbook, worksheet: Worksheet,
                     rules_config: CellRulesConfig):
        self.cli.print("[Info] Processing finished!", style=STYLE_INFO)
        self.cli.print("[Info] Saving...", style=STYLE_INFO)

        target_xlsx = os.path.normpath(os.path.abspath(self.target_xlsx))
        save_as = self.save_as
        overwrite = self.overwrite
        save_anyway = self.save_anyway

        if not save_as:
            save_as = target_xlsx
            overwrite = True

        dest = os.path.normpath(os.path.abspath(save_as))

        if not save_anyway and self._execution_err is not None:
            self.cli.print(f"[Warning] '{dest}' will not be saved because an error occurred during processing, "
                           f"use --save-anyway to ignore errors.", style=STYLE_WARN)
            return
        self.save_xlsx_file(workbook, dest=dest, overwrite=overwrite)
        self._execution_err = None
        self.disconnect_signals()

    def _before_executed(self, processor: CellRulesProcessor, context: FunctionContext,
                         rule_scope_target: PrimitiveDataType | CellRule | List[CellRule]):
        if isinstance(rule_scope_target, list):
            count = len(rule_scope_target)
        elif isinstance(rule_scope_target, CellRule) or is_primitive_data_type(rule_scope_target):
            count = 1
        else:
            count = 0
        self.cli.print(f"[Info] Executing: '{context.rule_scope}' with {count} rules", style=STYLE_INFO)

    def _on_succeed(self, processor: CellRulesProcessor, context: FunctionContext, rule: CellRule | PrimitiveDataType,
                    result: Any):
        self.cli.print(f"[Info] Executed: '{context.rule_scope}'", style=STYLE_SUCCESS)
        if self.verbose:
            self.cli.print(f"\tResult: {result}", style=STYLE_HINT)
            self.cli.print(f"\tRule: {rule}", style=STYLE_HINT)

    def _on_failed(self, processor: CellRulesProcessor, context: FunctionContext, rule: CellRule | PrimitiveDataType,
                   error: BaseException):
        self.cli.print(f"[Error] Failed: '{context.rule_scope}'", style=STYLE_ERROR)
        if self.verbose:
            self.cli.print(f"\tRule: {rule}", style=STYLE_HINT)
            self.cli.print_exception()
        else:
            self.cli.print(f"\tRule: {rule}", style=STYLE_HINT)
            self.cli.print(f"\tError: {error}", style=STYLE_HINT)

    def display_rules_config_info(self, rules_config: CellRulesConfig):
        self.cli.print(f"[Info] Total: {len(rules_config.rules)} scopes", style=STYLE_INFO)

    def get_workbook(self, target_xlsx: str, create: bool) -> Workbook | None:
        if create and not os.path.isfile(target_xlsx):
            self.cli.print(f"[Warning] '{target_xlsx}' not found", style=STYLE_WARN)
            self.cli.print(f"[Info] Creating file '{target_xlsx}'", style=STYLE_INFO)
            workbook = create_xlsx_file(path=target_xlsx)
            if os.path.isfile(target_xlsx):
                self.cli.print(f"[Info] '{target_xlsx}' created", style=STYLE_SUCCESS)
                return workbook
            else:
                self.cli.print(f"[Error] Failed to create '{target_xlsx}'", style=STYLE_ERROR)
                return None
        elif os.path.isfile(target_xlsx):
            try:
                workbook = load_workbook(target_xlsx)
            except BaseException as e:
                self.cli.print(f"[Error] Failed to load '{target_xlsx}'", style=STYLE_ERROR)
                if self.verbose:
                    self.cli.print_exception()
                else:
                    self.cli.print(f"\tError: {e}", style=STYLE_ERROR)
                return None
            else:
                return workbook
        else:
            self.cli.print(f"[Error] '{target_xlsx}' not found", style=STYLE_ERROR)
            return None

    def get_worksheet(self, workbook: Workbook, sheet: str | int | None) -> Worksheet | None:
        if sheet is None:
            return workbook.active
        elif isinstance(sheet, str):
            if sheet in workbook.sheetnames:
                return workbook[sheet]
            else:
                self.cli.print(f"[Error] Worksheet '{sheet}' not found", style=STYLE_WARN)
                self.cli.print(f"[Info] Creating worksheet '{sheet}'", style=STYLE_INFO)
                worksheet = workbook.create_sheet(title=sheet)
                self.cli.print(f"[Info] Worksheet '{sheet}' created", style=STYLE_SUCCESS)
                return worksheet
        else:
            try:
                worksheet = workbook.worksheets[sheet]
            except BaseException as e:
                self.cli.print(f"[Error] Worksheet '{sheet}' not found", style=STYLE_ERROR)
                if self.verbose:
                    self.cli.print_exception()
                else:
                    self.cli.print(f"\tError: {e}", style=STYLE_ERROR)
                return None
            else:
                return worksheet

    def get_rules_config(self, config_file: str, config_file_encoding: str) -> CellRulesConfig | None:
        if not os.path.exists(config_file):
            self.cli.print(f"[Error] '{config_file}' not found", style=STYLE_ERROR)
            return None
        try:
            rules_config = self.cli.rules_config_factory.load(config_file, encoding=config_file_encoding)
        except BaseException as e:
            self.cli.print(f"[Error] Failed to load rules config file '{config_file}'", style=STYLE_ERROR)
            if self.verbose:
                self.cli.print_exception()
            else:
                self.cli.print(f"\tError: {e}", style=STYLE_ERROR)
            return None
        else:
            return rules_config

    def check_save_as(self, save_as: str, overwrite: bool) -> bool:
        if os.path.exists(save_as) and not overwrite:
            self.cli.print(f"[Error] '{save_as}' already exists, use --overwrite to overwrite it",
                           style=STYLE_ERROR)
            return False
        return True

    def save_xlsx_file(self, workbook: Workbook, dest: str, overwrite: bool):
        if not self.check_save_as(dest, overwrite):
            return
        try:
            workbook.save(dest)
        except BaseException as e:
            self.cli.print(f"[Error] Failed to save to '{dest}'", style=STYLE_ERROR)
            if self.verbose:
                self.cli.print_exception()
            else:
                self.cli.print(f"\tError: {e}", style=STYLE_ERROR)
        else:
            self.cli.print(f"[Info] '{dest}' saved", style=STYLE_SUCCESS)

    def connect_signals(self):
        self.cli.rules_processor.started.connect(self._on_started)
        self.cli.rules_processor.finished.connect(self._on_finished)
        self.cli.rules_processor.before_executed.connect(self._before_executed)
        self.cli.rules_processor.succeed.connect(self._on_succeed)
        self.cli.rules_processor.failed.connect(self._on_failed)

    def disconnect_signals(self):
        self.cli.rules_processor.started.disconnect(self._on_started)
        self.cli.rules_processor.finished.disconnect(self._on_finished)
        self.cli.rules_processor.before_executed.disconnect(self._before_executed)
        self.cli.rules_processor.succeed.disconnect(self._on_succeed)
        self.cli.rules_processor.failed.disconnect(self._on_failed)

    def start(self, workbook: Workbook, worksheet: Worksheet, rules_config: CellRulesConfig, quick_fail: bool):
        if self.cli.rules_processor.is_processing():
            self.cli.print("[Error] Processing", style=STYLE_ERROR)
            return
        self.cli.rules_processor.process(
            workbook=workbook,
            worksheet=worksheet,
            rules_config=rules_config,
            quick_fail=quick_fail,
        )

    def start_async(self, workbook: Workbook, worksheet: Worksheet, rules_config: CellRulesConfig, quick_fail: bool):
        # 异步功能尚未实现，现阶段直接调用同步方法
        self.start(workbook, worksheet, rules_config, quick_fail)

    def run(self, *args, **kwargs):
        config_file = self.config_file
        config_file_encoding = self.config_file_encoding
        target_xlsx = os.path.normpath(os.path.abspath(self.target_xlsx))
        create = self.create
        sheet = self.sheet
        save_as = self.save_as
        run_async = self.run_async
        quick_fail = self.quick_fail
        overwrite = self.overwrite

        if not save_as:
            save_as = target_xlsx
            overwrite = True

        rules_config = self.get_rules_config(config_file, config_file_encoding)
        if not rules_config:
            return
        workbook = self.get_workbook(target_xlsx, create)
        if not workbook:
            return
        worksheet = self.get_worksheet(workbook, sheet)
        if not worksheet:
            return
        if not self.check_save_as(save_as, overwrite):
            return
        self.connect_signals()
        if not run_async:
            self.start(workbook, worksheet, rules_config, quick_fail)
        else:
            self.start(workbook, worksheet, rules_config, quick_fail)
