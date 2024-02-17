import threading
from typing import Any, Dict, Iterable, List

import blinker
from openpyxl.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from excelfiller.common import is_primitive_data_type, flatten, determine_values, get_cells_in_scope, PrimitiveDataType
from excelfiller.function import BaseFunction, FunctionContext, NULL
from excelfiller.functions.fakerfn import FakerFunctions
from excelfiller.functions.randomfn import RandomFunctions
from excelfiller.rule import CellRulesConfig, CellRule


class FunctionRegistry(object):
    def __init__(self):
        self._functions: Dict[str, BaseFunction] = {}
        self._lock = threading.Lock()

    def register(self, name: str, fn: BaseFunction):
        with self._lock:
            if name in self._functions:
                raise RuntimeError(f"function {name} is already registered")
            self._functions[name] = fn

    def get(self, name: str) -> BaseFunction | None:
        with self._lock:
            return self._functions.get(name, None)

    def unregister(self, name: str) -> BaseFunction | None:
        with self._lock:
            if name in self._functions:
                return self._functions.pop(name)
            return None

    def is_registered(self, name: str) -> bool:
        with self._lock:
            return self._functions.get(name, None) is not None

    def clear(self):
        with self._lock:
            self._functions.clear()


class CellRulesProcessor(object):
    BuiltinFunctions = {
        "random": RandomFunctions(),
        "fake": FakerFunctions(),
    }

    def __init__(self):
        self.function_registry = FunctionRegistry()
        self.processing_started = blinker.Signal()
        self.processing_finished = blinker.Signal()
        self.processing_errored = blinker.Signal()
        self.before_rule_executed = blinker.Signal()
        self.rule_execution_succeed = blinker.Signal()
        self.rule_execution_failed = blinker.Signal()

        self._worksheet_vars_map = {
            "MIN_COL": self._get_min_col,
            "MAX_COL": self._get_max_col,
            "MIN_ROW": self._get_min_row,
            "MAX_ROW": self._get_max_row,
        }

        self._stop_evnet: threading.Event = threading.Event()
        self._is_processing: bool = False

        self._register_builtin_functions()

    def process(self, workbook: Workbook, worksheet: Worksheet, rules_config: CellRulesConfig,
                quick_fail: bool = False):
        if self._is_processing:
            raise RuntimeError("a task is already running")
        self._is_processing = True
        # 发送started信号
        self.processing_started.send(self, workbook=workbook, worksheet=worksheet, rules_config=rules_config)
        # 调用主逻辑
        self._core_logic(rules_config, workbook, worksheet, stop_event=None, quick_fail=quick_fail)

    def process_async(self, workbook: Workbook, worksheet: Worksheet, rules_config: CellRulesConfig,
                      quick_fail: bool = False):
        if self._is_processing:
            raise RuntimeError("a task is already running")
        self._is_processing = True
        thread = threading.Thread(target=self._core_logic, kwargs={
            "rules_config": rules_config,
            "workbook": workbook,
            "worksheet": worksheet,
            "stop_event": self._stop_evnet,
            "quick_fail": quick_fail
        })
        # 发送started信号
        self.processing_started.send(self, workbook=workbook, worksheet=worksheet, rules_config=rules_config)
        thread.start()

    def is_processing(self):
        return self._is_processing

    def stop(self):
        self._stop_evnet.set()

    def _core_logic(self, rules_config: CellRulesConfig, workbook: Workbook, worksheet: Worksheet,
                    stop_event: threading.Event | None, quick_fail: bool = False):

        try:
            if stop_event:
                stop_event.clear()
            # 确定当前worksheet相关变量的具体值，如$MIN_COL、$MAX_COL、$MIN_ROW、$MAX_ROW等
            worksheet_vars = determine_values(self._worksheet_vars_map, worksheet)
        except BaseException as e:
            self.processing_errored.send(self, workbook=workbook, worksheet=worksheet, rules_config=rules_config,
                                         error=e)
            self._is_processing = False
            self.processing_finished.send(self, workbook=workbook, worksheet=worksheet, rules_config=rules_config)
            return

            # 从rules列表中取出每一个待处理的项
        for rule_scope, rule_scope_target in rules_config.rules.items():
            # 处理中途退出请求
            if stop_event and stop_event.is_set():
                break

            context = None
            cur_rule = None

            try:
                # 创建函数上下文
                context = FunctionContext(
                    rule_scope=rule_scope,
                    rules_globals=rules_config.rules_globals.copy(),
                    workbook=workbook,
                    worksheet=worksheet,
                    worksheet_vars=worksheet_vars.copy()
                )

                if is_primitive_data_type(rule_scope_target) or isinstance(rule_scope_target, CellRule):
                    # 如果rule_scope_target是单个的值及单个PrimitiveDataType或单个CellRule，则直接执行
                    cur_rule = rule_scope_target
                    # 发送before_executed信号
                    self.before_rule_executed.send(self, context=context, rule_scope_target=rule_scope_target)
                    self._execute_rule(context, cur_rule)
                elif isinstance(rule_scope_target, list):
                    # 发送before_executed信号
                    self.before_rule_executed.send(self, context=context, rule_scope_target=rule_scope_target)
                    # 如果rule_scope_target是CellRule列表，则执行列表中的每一个CellRule
                    for rule in rule_scope_target:
                        cur_rule = rule
                        self._execute_rule(context, cur_rule)
                else:
                    raise TypeError(f"invalid type of rule_scope_target: {type(rule_scope_target)}")
            except BaseException as e:
                # 发送failed信号
                self.rule_execution_failed.send(self, context=context, rule=cur_rule, error=e)
                # 如果quick_fail为True，则在发生错误时立即跳出循环
                if quick_fail:
                    break

        if stop_event:
            stop_event.clear()
        self._is_processing = False
        # 发送finished信号
        self.processing_finished.send(self, workbook=workbook, worksheet=worksheet, rules_config=rules_config)

    def _execute_rule(self, context: FunctionContext, rule_or_value: CellRule | PrimitiveDataType):
        """单条rule的执行逻辑"""
        if (not is_primitive_data_type(rule_or_value)) and (not isinstance(rule_or_value, CellRule)) and (
                rule_or_value != NULL):
            # 首先确保rule_or_value是CellRule或PrimitiveDataType
            raise TypeError(f"invalid type of rule: {type(rule_or_value)}")

        if rule_or_value == NULL:
            # 发送succeed信号
            self.rule_execution_succeed.send(self, context=context, rule=rule_or_value, result=NULL)
            return

        # 获取rule_scope中的全部cell
        cells = get_cells_in_scope(context.worksheet, context.rule_scope, context.worksheet_vars)
        if isinstance(cells, (Cell, MergedCell)):
            cells = [cells]

        # 如果rule_or_value是CellRule，则执行该CellRule
        if isinstance(rule_or_value, CellRule):
            fn = self.function_registry.get(rule_or_value.fn)
            if not isinstance(fn, BaseFunction):
                raise RuntimeError(f"function {rule_or_value.fn} is not registered")
            if rule_or_value.per_cell:
                # 如果指定了per_cell，则对cells中每个cell执行一次fn，并获取每次执行结果的列表
                result = self._execute_per_cell(fn, context, rule_or_value, cells)
            else:
                # 否则，对cells整体执行一次fn，即只执行一次fn
                result = self._execute_for_scope(fn, context, rule_or_value, cells)
            # 发送succeed信号
            self.rule_execution_succeed.send(self, context=context, rule=rule_or_value, result=result)
            return
        # 如果rule_or_value是PrimitiveDataType，则直接将其值赋给cells中的每个cell
        elif is_primitive_data_type(rule_or_value):
            result = rule_or_value
            for cell in flatten(cells):
                self._apply_result_to(cell, result)
            # 发送succeed信号
            self.rule_execution_succeed.send(self, context=context, rule=rule_or_value, result=result)
            return
        else:
            raise RuntimeError("LOGIC ERROR: IT SHOULD NEVER GOES HERE!")

    def _execute_per_cell(self, fn: BaseFunction, context: FunctionContext, rule: CellRule,
                          cells: Cell | MergedCell | Iterable) -> List[Any]:
        results = {}
        for current_cell in flatten(cells):
            # 对每个单元格执行fn
            result = fn.invoke(context, current_cell, *rule.args, **rule.kwargs)
            # 收集并记录执行结果
            results[current_cell] = result
        # 将执行结果运用到对应的单元格上
        for current_cell, cell_value in results.items():
            self._apply_result_to(current_cell, cell_value)
        return [*results.values()]

    def _execute_for_scope(self, fn: BaseFunction, context: FunctionContext, rule: CellRule,
                           cells: Cell | MergedCell | Iterable) -> Any:
        current_cell = None
        # 对整个scope执行fn，并收集执行结果
        result = fn.invoke(context, current_cell, *rule.args, **rule.kwargs)
        if result != NULL:
            # 将执行结果运用到scope内所有单元格上
            for cell in flatten(cells):
                self._apply_result_to(cell, result)
        return result

    @staticmethod
    def _apply_result_to(cell: Cell | MergedCell, result: Any):
        if result == NULL:
            return
        if isinstance(cell, MergedCell):
            return
        if isinstance(cell, Cell):
            cell.value = result
            return
        raise TypeError(f"invalid type of cell: {type(cell)}")

    def _register_builtin_functions(self):
        for name, fn in self.BuiltinFunctions.items():
            if not self.function_registry.is_registered(name):
                self.function_registry.register(name, fn)

    @staticmethod
    def _get_max_row(worksheet: Worksheet) -> int:
        return worksheet.max_row

    @staticmethod
    def _get_max_col(worksheet: Worksheet) -> str:
        return get_column_letter(worksheet.max_column)

    @staticmethod
    def _get_min_row(worksheet: Worksheet) -> int:
        return worksheet.min_row

    @staticmethod
    def _get_min_col(worksheet: Worksheet) -> str:
        return get_column_letter(worksheet.min_column)
